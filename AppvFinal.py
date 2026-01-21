from __future__ import annotations
# ---- Silence "missing ScriptRunContext" warnings from Streamlit worker threads
import logging, warnings



# Cover the whole Streamlit logger tree (some versions use parent loggers)
for name in (
    "streamlit.runtime.scriptrunner.script_run_context",
    "streamlit.runtime.scriptrunner",
    "streamlit.runtime",
    "streamlit",
):
    logger = logging.getLogger(name)
    logger.setLevel(logging.ERROR)
    # Optional: in case Streamlit pre-attached handlers at WARN level
    for h in list(logger.handlers):
        h.setLevel(logging.ERROR)

# (Optional) If any variants are emitted via Python warnings instead of logging
warnings.filterwarnings("ignore", message=r".*missing ScriptRunContext!.*")


# =========================
# G3M Ultime — Optimized 10x (Fixed)
# =========================
import io
import os
import re
import hashlib
import platform
import inspect
from pathlib import Path
from functools import lru_cache
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import pandas as pd

# ---- Global perf flags
pd.set_option("mode.copy_on_write", True)  # pandas 2.x: avoids hidden copies
np.set_printoptions(legacy="1.25")  # keep numpy formatting stable if needed

import streamlit as st
# Attach ScriptRunContext to worker threads so Streamlit won't warn
try:
    from streamlit.runtime.scriptrunner import add_script_run_ctx, get_script_run_ctx
    _MAIN_CTX = get_script_run_ctx()
    def submit_with_ctx(pool: ThreadPoolExecutor, fn, *args, **kwargs):
        import threading
        def _runner():
            try:
                if _MAIN_CTX is not None:
                    add_script_run_ctx(thread=threading.current_thread(), ctx=_MAIN_CTX)
            except Exception:
                pass
            return fn(*args, **kwargs)
        return pool.submit(_runner)
except Exception:
    # Fallback: if API not available, just use pool.submit (logging suppression above still helps)
    def submit_with_ctx(pool: ThreadPoolExecutor, fn, *args, **kwargs):
        return pool.submit(fn, *args, **kwargs)

# -----------------------------------
# STREAMLIT: Cache & resource policies
# -----------------------------------
@st.cache_data(show_spinner=False)
def _cached_cached_ultime_cycle(df_orig: pd.DataFrame, method_names: list[str], start_year: int, *, decay_r: float = 0.60):
    # Cached wrapper to avoid recomputing identical heavy runs
    return _cached_ultime_cycle(df_orig, method_names, start_year, decay_r=decay_r)

@st.cache_resource(show_spinner=False)
def _get_threadpool(max_workers: int | None = None) -> ThreadPoolExecutor:
    # Reuse a single threadpool across reruns
    if max_workers is None:
        cpu = os.cpu_count() or 4
        max_workers = max(2, min(8, cpu))  # sensible default
    return ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="g3m-exec")

# =========================
# EXTERNAL IMPORTS (if any)
# =========================
try:
    from FinalHypotheseCadence import AssumptionCalendar
except Exception:
    AssumptionCalendar = None

try:
    from FinalHypotheseCorrFact import AssumptionFactors
except Exception:
    AssumptionFactors = None

try:
    # Keep names identical to preserve logic/compatibility
    from AppMethods import (
        chainladder,
        vylder,
        combinaison,
        Poissonbootstrap,   # stays available from AppMethods
        moyenne_facteurs,
        ChainLadderBootstrap,
        dernier_fact,
        last3,
        #BornhuetterFerguson,
    )
except Exception:
    chainladder = None
    vylder = None
    combinaison = None
    Poissonbootstrap = None
    moyenne_facteurs = None
    ChainLadderBootstrap = None
    dernier_fact = None
    last3 = None
    #BornhuetterFerguson = None

st.set_page_config(page_title="Triangle Completion • Ultime (Optimized)", layout="wide")

# Only keep methods that may be available
ALL_METHOD_NAMES = [
    "last3",
    "chainladder",
    "moyenne_facteurs",
    "vylder",
    "combinaison",
    #"BornhuetterFerguson",
    "Poissonbootstrap",
    "ChainLadderBootstrap",
    "dernier_fact"
]

# Friendly names & descriptions (edit freely)
METHOD_META = {
    "last3": {"title": "Last 3", "desc": "Uses the last three observed development ratios to project ultimates."},
    "chainladder": {"title": "Chain Ladder (Deterministic)", "desc": "Classical deterministic chain ladder using age-to-age factors."},
    "moyenne_facteurs": {"title": "Average Factors", "desc": "Completes using average development factors (mean-based)."},
    "vylder": {"title": "De Vylder", "desc": "Least-squares method minimizing squared deviations of development factors."},
    "combinaison": {"title": "Combination", "desc": "Blend of multiple deterministic methods for robustness."},
    #"BornhuetterFerguson": {"title": "Bornhuetter-Ferguson", "desc": "Uses loss ratio (Ultimate/Premium from first year) to project ultimates for all years."},
    "Poissonbootstrap": {"title": "Poisson Bootstrap (ODP)", "desc": "ODP bootstrap on incremental triangle; simulates completion distribution."},
    "ChainLadderBootstrap": {"title": "Chain Ladder Bootstrap", "desc": "Bootstrap around chain ladder with resampling of residuals."},
    "dernier_fact": {"title": "Last Factor", "desc": "Projects using the last observed factor per development age."}
}

# =========================
# THEME / CSS
# =========================
def _init_theme_state():
    """Initialize theme state on first run, using system preference if available."""
    if "theme_mode" not in st.session_state:
        # Inject JavaScript to detect Windows/browser system preference
        st.markdown(
            """
            <script>
            (function() {
              // Check browser's computed preference (works with Windows dark mode)
              const prefersDark = window.matchMedia('(prefers-color-scheme: dark)').matches;
              const prefersLight = window.matchMedia('(prefers-color-scheme: light)').matches;
              
              // Get or set localStorage key
              const storageKey = 'g3m_theme_mode';
              let theme = localStorage.getItem(storageKey);
              
              // If first time or no saved preference, use system preference
              if (!theme) {
                theme = prefersDark ? 'dark' : 'light';
              }
              
              // Always save to localStorage for persistence
              localStorage.setItem(storageKey, theme);
              
              // Store globally for Streamlit to pick up
              window.g3m_system_theme = theme;
              
              // Also try to inject via data attribute
              document.documentElement.setAttribute('data-g3m-theme', theme);
            })();
            </script>
            """,
            unsafe_allow_html=True,
        )
        
        # Try to detect from JavaScript, default to light if dark mode not detected
        # This will pick up the theme from the script above
        st.session_state.theme_mode = "dark"  # Will be overridden by JS below if light mode is system default

def _get_theme_colors(mode: str) -> dict:
    """Return color palette based on theme mode."""
    if mode == "light":
        return {
            "primary": "#2F7ED8",
            "accent": "#F4A300",
            "bg": "#F4F7FB",
            "surface": "#FFFFFF",
            "surface_alt": "#F8FAFC",
            "text": "#1E293B",
            "muted": "#6B7280",
            "border": "#E2E8F0",
            "border_input": "#E5E7EB",
            "success": "#3CB371",
            "warning": "#E6B800",
            "error": "#E5533D",
            "table_header_bg": "rgba(47,126,216,0.10)",
            "table_header_text": "#2F7ED8",
            "table_alt_row": "#F8FAFC",
            "sidebar_border": "#E8EEF7",
        }
    else:  # dark mode
        return {
            "primary": "#2F7ED8",
            "accent": "#F4A300",
            "bg": "#0B1320",
            "surface": "#111A2E",
            "surface_alt": "#0E172A",
            "text": "#EAF2FF",
            "muted": "#9FB0C8",
            "border": "#1F2A40",
            "border_input": "#1F2A40",
            "success": "#3CB371",
            "warning": "#E6B800",
            "error": "#E5533D",
            "table_header_bg": "rgba(47,126,216,0.12)",
            "table_header_text": "#2F7ED8",
            "table_alt_row": "#0E172A",
        }

def _inject_brand_css(colors: dict):
    """Inject theme CSS with dynamic colors."""
    st.markdown(
        f"""
        <style>
        :root {{
          --g3m-primary:{colors['primary']};
          --g3m-accent:{colors['accent']};
          --g3m-bg:{colors['bg']};
          --g3m-surface:{colors['surface']};
          --g3m-surface-alt:{colors['surface_alt']};
          --g3m-text:{colors['text']};
          --g3m-muted:{colors['muted']};
          --g3m-border:{colors['border']};
          --g3m-border-input:{colors.get('border_input', colors['border'])};
          --g3m-success:{colors['success']};
          --g3m-warning:{colors['warning']};
          --g3m-error:{colors['error']};
          --g3m-table-header-bg:{colors.get('table_header_bg', 'rgba(47,126,216,0.12)')};
          --g3m-table-header-text:{colors.get('table_header_text', '#2F7ED8')};
          --g3m-table-alt-row:{colors.get('table_alt_row', '#0E172A')};
          --g3m-sidebar-border:{colors.get('sidebar_border', '#E8EEF7')};
        }}
        
        html, body, .stApp, section.main, .block-container {{
          background: var(--g3m-bg) !important;
          color: var(--g3m-text) !important;
        }}
        
        /* Base text color – scoped, safe (not global) */
        .stMarkdown, 
        .stText, 
        .stCaption,
        .stHeader,
        .stSubheader,
        label,
        p {{
          color: var(--g3m-text);
        }}
        
        [data-testid="stSidebar"] > div:first-child {{
          background: var(--g3m-surface) !important;
          border-right: 1px solid var(--g3m-sidebar-border) !important;
        }}
        
        /* Phase structure with improved vertical rhythm */
        .g3m-phase {{
          margin-bottom: 2.5rem;
          margin-top: 0.5rem;
        }}
        
        .g3m-phase h3 {{
          margin-top: 1.5rem;
          margin-bottom: 0.25rem;
          letter-spacing: 0.01em;
        }}
        
        /* Cards: intentional, layered appearance */
        .g3m-card {{
          padding: 1.5rem;
          background: var(--g3m-surface);
          border: 1px solid var(--g3m-border);
          border-left: 5px solid var(--g3m-accent);
          border-radius: 12px;
          box-shadow: 0 2px 8px rgba(0,0,0,0.04);
          margin-bottom: 1.25rem;
        }}
        
        .g3m-card strong {{
          color: var(--g3m-primary);
        }}
        
        /* Inputs: embedded, not floating */
        .stTextInput input, .stNumberInput input,
        .stTextArea textarea,
        .stDateInput input, .stTimeInput input {{
          background: var(--g3m-surface) !important;
          border: 1px solid #E5E7EB !important;
          border-radius: 8px !important;
          color: var(--g3m-text) !important;
          font-size: 0.95rem !important;
          padding: 0.75rem 0.875rem !important;
        }}
        
        /* Selectbox containers - style only the wrapper, not the input */
        .stSelectbox div[data-baseweb="select"] > div,
        .stMultiSelect div[data-baseweb="select"] > div {{
          background: var(--g3m-surface) !important;
          border: 1px solid #E5E7EB !important;
          border-radius: 8px !important;
          padding: 0 !important;
        }}
        
        .stTextInput input:focus, .stNumberInput input:focus,
        .stTextArea textarea:focus,
        .stDateInput input:focus, .stTimeInput input:focus,
        .stSelectbox div[data-baseweb="select"] > div:focus,
        .stMultiSelect div[data-baseweb="select"] > div:focus {{
          border-color: var(--g3m-primary) !important;
          box-shadow: none !important;
          outline: none !important;
        }}
        
        input::placeholder, textarea::placeholder {{ 
          color: #A0AEC0 !important; 
          opacity: 1 !important; 
        }}
        
        /* FIX: make selected value visible in selectbox (light & dark) */
        [data-baseweb="select"] input {{
          color: var(--g3m-text) !important;
          background-color: transparent !important;
          z-index: 2 !important;
          padding: 0.75rem 0.875rem !important;
          font-size: 0.95rem !important;
        }}
        
        /* Target the text content directly in BaseWeb select */
        [data-baseweb="select"] {{
          color: var(--g3m-text) !important;
        }}
        
        [data-baseweb="select"] * {{
          color: var(--g3m-text) !important;
        }}
        
        .stSelectbox [data-baseweb="select"] {{
          color: var(--g3m-text) !important;
        }}
        
        .stSelectbox [data-baseweb="select"] input {{
          color: var(--g3m-text) !important;
          caret-color: var(--g3m-text) !important;
          padding: 0.75rem 0.875rem !important;
          font-size: 0.95rem !important;
        }}
        
        .stSelectbox [data-baseweb="select"] * {{
          color: var(--g3m-text) !important;
        }}
        
        /* Force visible text in select value display */
        .stSelectbox input {{
          color: var(--g3m-text) !important;
          caret-color: var(--g3m-text) !important;
        }}
        
        /* Placeholder style in selectbox */
        [data-baseweb="select"] input::placeholder {{
          color: var(--g3m-muted) !important;
          opacity: 1 !important;
        }}
        
        /* Dropdown menu styling */
        [data-baseweb="menu"] {{
          background: var(--g3m-surface) !important;
        }}
        
        [data-baseweb="menu"] li {{
          color: var(--g3m-text) !important;
        }}
        
        [data-baseweb="menu"] li:hover {{
          background: rgba(47,126,216,0.10) !important;
          color: var(--g3m-text) !important;
        }}
        
        input[type="radio"], input[type="checkbox"], input[type="range"] {{ 
          accent-color: var(--g3m-accent) !important; 
        }}
        
        /* Buttons: professional, intentional */
        .stButton>button[kind="primary"] {{
          background: var(--g3m-primary) !important;
          color: #FFFFFF !important;
          font-weight: 700;
          border-radius: 8px;
          text-shadow: none;
          border: none;
          transition: none;
          padding: 0.625rem 1.25rem !important;
        }}
        
        .stButton>button[kind="primary"]:hover {{
          filter: brightness(1.05);
          background: var(--g3m-primary) !important;
          color: #FFFFFF !important;
        }}
        
        .stDownloadButton>button {{
          background: var(--g3m-accent) !important;
          color: #1E293B !important;
          font-weight: 700;
          border-radius: 8px;
          text-shadow: none;
          border: none;
          transition: none;
          padding: 0.625rem 1.25rem !important;
        }}
        
        .stDownloadButton>button:hover {{
          filter: brightness(1.04);
          background: var(--g3m-accent) !important;
          color: #1E293B !important;
        }}
        
        /* File upload: premium drop zone */
        [data-testid="stFileUploaderDropzone"] {{
          background: var(--g3m-surface-alt) !important;
          border: 1px solid #E5E7EB !important;
          border-radius: 12px !important;
          padding: 2rem 1rem !important;
        }}
        
        [data-testid="stFileUploader"] button {{
          background: var(--g3m-primary) !important;
          border-color: var(--g3m-primary) !important;
          color: white !important;
          border-radius: 8px !important;
          font-weight: 700 !important;
          padding: 0.625rem 1rem !important;
        }}
        
        /* Tags in multiselect */
        .stMultiSelect [data-baseweb="tag"] {{
          background: rgba(47, 126, 216, 0.12) !important;
          border-color: rgba(47, 126, 216, 0.30) !important;
          color: var(--g3m-text) !important;
        }}
        
        /* Tables: Excel-like, professional appearance */
        .stDataFrame div[role="table"] {{
          background: var(--g3m-surface) !important;
          font-size: 0.9rem !important;
        }}
        
        .stDataFrame div[role="table"] thead th {{
          background: var(--g3m-table-header-bg) !important;
          color: var(--g3m-table-header-text) !important;
          font-weight: 700 !important;
          border-bottom: 2px solid var(--g3m-primary) !important;
          border-right: 1px solid #E5E7EB !important;
          padding: 0.875rem 0.75rem !important;
          text-align: left;
        }}
        
        .stDataFrame div[role="table"] tbody tr {{
          border-bottom: 1px solid #F0F0F0 !important;
          height: 2.5rem;
        }}
        
        .stDataFrame div[role="table"] tbody tr:nth-child(even) {{
          background: var(--g3m-table-alt-row) !important;
        }}
        
        .stDataFrame div[role="table"] tbody td {{
          border-right: 1px solid #F0F0F0 !important;
          padding: 0.875rem 0.75rem !important;
          color: var(--g3m-text) !important;
        }}
        
        /* Typography and semantic colors */
        h1 {{ 
          color: var(--g3m-primary) !important;
          font-weight: 900 !important;
          letter-spacing: 0.01em;
        }}
        
        h2, h3, h4 {{ 
          color: var(--g3m-primary) !important;
          font-weight: 700 !important;
        }}
        
        .g3m-muted {{ 
          color: #9CA3AF !important; 
          font-size: 0.9rem !important;
        }}
        
        .g3m-success {{ 
          color: var(--g3m-success) !important; 
          font-weight: 600;
        }}
        
        .g3m-warning {{ 
          color: var(--g3m-warning) !important; 
          font-weight: 600;
        }}
        
        .g3m-error {{ 
          color: var(--g3m-error) !important; 
          font-weight: 600;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

def _brand_header():
    _init_theme_state()
    
    # Inject JavaScript to read system theme from localStorage and apply it
    script = """
    <script>
    (function() {
      const storageKey = 'g3m_theme_mode';
      const prefersDark = window.matchMedia('(prefers-color-scheme: dark)').matches;
      
      // Read from localStorage, or detect from system preference
      let theme = localStorage.getItem(storageKey);
      if (!theme) {
        theme = prefersDark ? 'dark' : 'light';
        localStorage.setItem(storageKey, theme);
      }
      
      // Store for Python to access on next rerun
      window.g3m_system_theme = theme;
    })();
    </script>
    """
    st.markdown(script, unsafe_allow_html=True)
    
    colors = _get_theme_colors(st.session_state.theme_mode)
    _inject_brand_css(colors)
    
    left, center, right = st.columns([1, 3, 1])
    
    with left:
        st.image("g3m.png", width=150)
    
    with center:
        st.markdown(
            f"""
            <div style="font-weight:900;font-size:1.6rem;color:{colors['text']};letter-spacing:.02em;">G3M <span style="color:{colors['accent']};">CONSULTING</span></div>
            <div style="color:{colors['muted']};font-size:0.85rem;">Mesurer vos risques…Minimiser leurs impacts</div>
            """,
            unsafe_allow_html=True,
        )
    
    with right:
        def _toggle_theme():
            st.session_state.theme_mode = "light" if st.session_state.theme_mode == "dark" else "dark"
        
        theme_icon = "🌙" if st.session_state.theme_mode == "dark" else "☀️"
        
        st.button(theme_icon, key="theme_toggle", on_click=_toggle_theme)

# =========================
# DATA COERCION (Optimized)
# =========================
def _coerce_index_to_int(df: pd.DataFrame) -> pd.DataFrame:
    if np.issubdtype(df.index.dtype, np.integer):
        return df
    idx_str = df.index.astype(str).str.strip()
    idx_num = pd.to_numeric(idx_str, errors="coerce")
    if idx_num.isna().all():
        raise ValueError("Index could not be coerced to integers (accident years).")
    out = df.loc[~idx_num.isna()]
    out.index = np.rint(idx_num[~idx_num.isna()]).astype(int)

    out.sort_index(inplace=True)
    return out

def _coerce_age_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Keep only columns that can be int ages, preserve order by numeric value
    keep_map = {}
    for c in df.columns:
        try:
            keep_map[c] = int(str(c).strip())
        except Exception:
            continue
    if not keep_map:
        raise ValueError("No integer-like development-age columns found.")
    cols_sorted = [k for k,_ in sorted(keep_map.items(), key=lambda x: x[1])]
    sub = df.loc[:, cols_sorted]
    sub.columns = [int(round(float(str(c).strip()))) for c in sub.columns]

    # Numeric coercion once, reuse
    return sub.apply(pd.to_numeric, errors="coerce")

# ---------- Fast utilities on CLEAN triangles only ----------
def _cut_triangle_fast(df_clean: pd.DataFrame, cut_year: int) -> pd.DataFrame:
    # df_clean: index=int AY, columns=int ages (1..N), numeric/NaN
    first_year = int(df_clean.index.min())
    N = int(cut_year - first_year + 1)
    if N < 1:
        raise ValueError(f"cut_year {cut_year} is earlier than first accident year {first_year}.")
    df_rows = df_clean.loc[df_clean.index <= cut_year]
    ages_target = list(range(1, N + 1))
    df_cols = df_rows.reindex(columns=ages_target)
    ay = df_cols.index.to_numpy(dtype=int)[:, None]
    ages = np.arange(1, N + 1, dtype=int)[None, :]
    max_age_allowed = (cut_year - ay + 1)
    mask = (ages <= max_age_allowed)
    return df_cols.where(mask)

def _ultimes_from_completed_fast(completed: pd.DataFrame) -> pd.Series:
    # completed already clean/aligned
    return completed.iloc[:, -1].rename("Ultimate")

# =========================
# CACHING HASH HELPERS
# =========================
def _hash_frame_shape_values(df: pd.DataFrame) -> str:
    s = f"{df.shape[0]}x{df.shape[1]}"
    nnz = np.count_nonzero(~pd.isna(df.to_numpy(dtype=float, copy=False)))
    return f"{s}:{nnz}"

@lru_cache(maxsize=1024)
def _memo_cut_key(first_year: int, last_year: int, cut_year: int) -> tuple:
    return (first_year, last_year, cut_year)

# Cache method outputs per (method, cut_key, light hash)
_method_cut_cache: dict[tuple[str, tuple, str], pd.DataFrame] = {}

def _cache_get_method(method_name: str, cut_key: tuple, tri_hash: str) -> pd.DataFrame | None:
    return _method_cut_cache.get((method_name, cut_key, tri_hash))

def _cache_put_method(method_name: str, cut_key: tuple, tri_hash: str, df: pd.DataFrame) -> None:
    _method_cut_cache[(method_name, cut_key, tri_hash)] = df.copy(deep=True)

# =========================
# METHOD RUNNERS (Parallel)
# =========================
def _make_writable_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Return a DataFrame backed by a C-contiguous, writable float64 ndarray.
    This prevents 'assignment destination is read-only' when external methods
    mutate df.values in place.
    """
    arr = np.ascontiguousarray(df.to_numpy(dtype=float, na_value=np.nan), dtype=float)
    try:
        arr.setflags(write=True)
    except Exception:
        pass
    return pd.DataFrame(arr, index=df.index.copy(), columns=df.columns.copy())

def _ensure_incremental(df_cum: pd.DataFrame) -> pd.DataFrame:
    """Convert cumulative to incremental while preserving NaN boundaries."""
    inc = df_cum.copy()
    if inc.shape[1] >= 2:
        # Get numpy arrays
        arr = inc.to_numpy()
        
        # Compute differences
        diff = arr[:, 1:] - arr[:, :-1]
        
        # Apply differences
        result = arr.copy()
        result[:, 1:] = diff
        
        # Create mask: NaN where original was NaN OR previous was NaN
        na_mask = np.isnan(arr)
        
        # For columns 1+, also mask if previous column was NaN
        for j in range(1, arr.shape[1]):
            na_mask[:, j] = na_mask[:, j] | na_mask[:, j-1]
        
        # Apply mask
        result[na_mask] = np.nan
        
        # Return as DataFrame
        return pd.DataFrame(result, index=inc.index, columns=inc.columns)
    
    return inc

def _call_with_signature(func, df_in: pd.DataFrame, params: dict | None = None) -> pd.DataFrame:
    """
    Call `func(df_in, **filtered_kwargs)` where kwargs are filtered to only those
    present in the function's signature. Avoids 'unexpected keyword argument' errors.
    """
    params = params or {}
    try:
        sig = inspect.signature(func)
        allowed = {k: v for k, v in params.items() if k in sig.parameters}
    except (ValueError, TypeError):
        allowed = {}
    return func(df_in, **allowed)

def _call_method(func, df_cut: pd.DataFrame, name: str | None = None) -> pd.DataFrame:
    """
    Call an external completion method on a CLEAN triangle.

    Guarantees:
    - Always returns a pandas DataFrame (completed triangle)
    - Correctly handles bootstrap methods returning (df, bytes)
    - Stores bootstrap Excel bytes in st.session_state["BOOTSTRAP_REPORTS"]
    - Supplies a writable float64 frame
    - Filters kwargs to only what the target function supports
    """

    # --- Prepare writable input
    df_in = _make_writable_df(df_cut)

    # Poisson bootstrap expects incremental triangle
    if name == "Poissonbootstrap":
        df_in = _ensure_incremental(df_in)
        if df_in.shape[0] > 0 and pd.notna(df_in.iloc[0, 0]) and df_in.iloc[0, 1:].isna().all():
            # First AY has only first column - ensure it's not lost
            # Store the value to verify it's preserved
            first_ay_first_col = df_in.iloc[0, 0]
    

    # --- Fetch method parameters
    cfgs = st.session_state.get("METHOD_CONFIGS", {})
    params = dict(cfgs.get(name or "", {})) if isinstance(cfgs, dict) else {}

    # Chain Ladder Bootstrap default
    if name == "ChainLadderBootstrap":
        params.setdefault("Mode", "Med")
    
    # Bornhuetter-Ferguson needs primes data
    if name == "BornhuetterFerguson":
        primes_data = st.session_state.get("PRIMES_DATA", None)
        if primes_data is not None:
            params["primes_df"] = primes_data

    # --- Call method safely
    with pd.option_context("mode.copy_on_write", False):
        try:
            out = _call_with_signature(func, df_in, params)
        except (ValueError, TypeError) as e:
            msg = str(e).lower()
            if "read-only" in msg or "assignment destination is read-only" in msg:
                out = _call_with_signature(func, _make_writable_df(df_cut), params)
            else:
                raise

    # --- Handle bootstrap tuple return
    if isinstance(out, tuple):
        out_df, report_bytes = out

        if report_bytes is not None:
            st.session_state.setdefault("BOOTSTRAP_REPORTS", {})
            st.session_state.BOOTSTRAP_REPORTS[name] = report_bytes
    else:
        out_df = out

    # --- Validate output
    if not isinstance(out_df, pd.DataFrame):
        raise TypeError(f"Method '{name}' did not return a DataFrame")

    # --- Normalize index & columns
    return _coerce_age_columns(_coerce_index_to_int(out_df))

def _run_methods_parallel(df_cut: pd.DataFrame, method_names: list[str]) -> dict[str, pd.DataFrame]:
    results: dict[str, pd.DataFrame] = {}
    if not method_names:
        return results

    g = globals()
    work: list[tuple[str, callable]] = []
    for name in method_names:
        func = g.get(name, None)
        if callable(func):
            work.append((name, func))
        else:
            st.warning(f"Method '{name}' not found; skipping.")

    if not work:
        return results

    pool = _get_threadpool()
    futures = {submit_with_ctx(pool, _call_method, func, df_cut, name): name for name, func in work}

    for fut in as_completed(futures):
        name = futures[fut]
        try:
            results[name] = fut.result()
        except Exception as e:
            st.error(f"Method '{name}' failed: {e}")

    return results

# =========================
# ULTIMES COLLECTION
# =========================
def _complete_all_and_collect_ultimes(df_cut: pd.DataFrame, method_names: list[str], *, cut_key: tuple) -> tuple[dict[str, pd.DataFrame], pd.DataFrame]:
    tri_hash = _hash_frame_shape_values(df_cut)

    # Try cached results
    cached_preds: dict[str, pd.DataFrame] = {}
    remaining: list[str] = []
    for m in method_names:
        hit = _cache_get_method(m, cut_key, tri_hash)
        if isinstance(hit, pd.DataFrame):
            cached_preds[m] = hit
        else:
            remaining.append(m)

    preds = cached_preds.copy()
    if remaining:
        fresh = _run_methods_parallel(df_cut, remaining)
        for m, tri in fresh.items():
            _cache_put_method(m, cut_key, tri_hash, tri)
        preds.update(fresh)

    # Build ultimes table
    ult_tab = {}
    for m, tri in preds.items():
        try:
            ult_tab[m] = _ultimes_from_completed_fast(tri)
        except Exception:
            continue
    ult_df = pd.DataFrame(ult_tab) if ult_tab else pd.DataFrame()
    return preds, ult_df

# =========================
# ULTIME ENGINE (Optimized)
# =========================
def _cached_ultime_cycle(df_orig: pd.DataFrame, method_names: list[str], start_year: int, *, decay_r: float = 0.60):
    """Run cuts from start_year..last_year; collect per-cut ultimes for all methods."""
    full = _coerce_age_columns(_coerce_index_to_int(df_orig))
    years = full.index.to_numpy(dtype=int, copy=False)
    first_year = int(years.min())
    last_year = int(years.max())
    if not (first_year <= start_year < last_year):
        raise ValueError(f"Start year must be between {first_year} and {last_year-1}.")

    cuts = list(range(start_year, last_year + 1))
    sequences: dict[str, dict[int, list[float]]] = {m: {int(ay): [] for ay in years.tolist()} for m in method_names}
    artifacts: dict[int, dict] = {}
    final_ultimes = pd.DataFrame()

    first_year_int, last_year_int = int(first_year), int(last_year)

    for cut in cuts:
        cut_key = _memo_cut_key(first_year_int, last_year_int, int(cut))
        df_cut = _cut_triangle_fast(full, cut)
        preds, ult_df = _complete_all_and_collect_ultimes(df_cut, method_names, cut_key=cut_key)
        artifacts[cut] = {"cut_df": df_cut, "preds": preds, "ultimes_df": ult_df}

        if not ult_df.empty:
            for m, col in ult_df.items():
                for ay, val in col.items():
                    ay_int = int(np.rint(float(ay)))
                    sequences[m][ay_int].append(float(val) if pd.notna(val) else np.nan)

        if cut == cuts[-1]:
            final_ultimes = ult_df

    return sequences, artifacts, final_ultimes, cuts

# =========================
# BEST-METHOD SELECTION
# =========================
def _select_best_method(
    sequences: dict[str, dict[int, list[float]]],
    cuts: list[int],
    chosen_ay: int,
    *,
    closeness_weight: float = 0.70,  # closeness to final ultime
    decay_r: float = 0.85,          # heavier weight on later cuts for closeness
) -> tuple[str | None, pd.DataFrame]:
    """
    score = w * (weighted MAE to final ultime across cuts for chosen_ay)
          + (1 - w) * (std of ultimes across cuts for chosen_ay)
    Lower is better.
    """
    methods = list(sequences.keys())
    T = len(cuts)
    if T < 2:
        return None, pd.DataFrame()

    # weights for closeness: later cuts get higher weights
    w = np.array([decay_r ** (T - 2 - t) for t in range(T - 1)], dtype=float)
    w = w / w.sum() if w.sum() > 0 else np.ones(T - 1) / (T - 1)

    rows = []
    for m in methods:
        seq = sequences[m].get(int(chosen_ay), [])
        if len(seq) < 2 or all(pd.isna(x) for x in seq):
            rows.append((m, np.inf, np.inf, np.inf))
            continue
        U_final = seq[-1]
        hist = np.asarray(seq[:-1], dtype=float)
        errs = np.abs(hist - U_final)

        if np.isnan(U_final) or (errs.size == 0) or np.all(np.isnan(errs)):
            rows.append((m, np.inf, np.inf, np.inf))
            continue

        mask = ~np.isnan(errs)
        if mask.any():
            ww = w[:mask.size]
            ww = ww / ww.sum() if ww.sum() > 0 else np.ones(mask.sum()) / mask.sum()
            closeness = float(np.nansum(ww * errs))
        else:
            closeness = np.inf

        vol = float(np.nanstd(hist, ddof=0)) if np.isfinite(hist).any() else np.inf
        score = closeness_weight * closeness + (1 - closeness_weight) * vol
        rows.append((m, closeness, vol, score))

    df = pd.DataFrame(rows, columns=["method", "closeness_to_final", "stability_vol", "score"]).sort_values(
        ["score", "closeness_to_final", "stability_vol"], kind="mergesort"
    )
    best = None if df.empty or not np.isfinite(df["score"].iloc[0]) else str(df["method"].iloc[0])
    return best, df

# =========================
# FORMATTING HELPERS
# =========================
def _ceil_numeric(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    for c in out.columns:
        if pd.api.types.is_numeric_dtype(out[c]):
            s = pd.to_numeric(out[c], errors="coerce")
            s = s.replace([np.inf, -np.inf], np.nan)
            out[c] = np.rint(np.ceil(s)).astype("Int64")
    return out

def _format_display(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()

    def _fmt_series(s: pd.Series) -> pd.Series:
        if not pd.api.types.is_numeric_dtype(s):
            return s
        v = pd.to_numeric(s, errors="coerce")
        vals = v.to_numpy()
        finite = np.isfinite(vals)

        txt = pd.Series([""] * len(s), index=s.index, dtype="object")

        if finite.any():
            vf = np.rint(np.ceil(v[finite])).astype("Int64")
            txt.loc[finite] = vf.apply(lambda x: "" if pd.isna(x) else f"{int(x):,}".replace(",", " "))

        pos_inf = ~finite & (vals > 0)
        neg_inf = ~finite & (vals < 0)
        if pos_inf.any():
            txt.iloc[np.where(pos_inf)[0]] = "∞"
        if neg_inf.any():
            txt.iloc[np.where(neg_inf)[0]] = "-∞"

        return txt

    for c in out.columns:
        out[c] = _fmt_series(out[c])
    return out

def _sanitize_filename(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", (name or "").strip())
    cleaned = cleaned.rstrip(" .")
    return cleaned or "Ultime_Report"

# =========================
# EXCEL STYLES (Gray/Orange/Dark-Blue — optimized)
# =========================
def _apply_global_excel_styles(workbook, titles: dict[str, str]):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    DARK_BLUE   = "183B66"
    LIGHT_TEXT  = "FFFFFF"
    COOL_GRAY   = "CFD6DE"
    HEADER_TXT  = "183B66"
    BORDER_COL  = "B7C0CD"
    ALT_ROW     = "EEF2F7"
    ORANGE_ACC  = "F4A300"

    header_fill = PatternFill(start_color=COOL_GRAY, end_color=COOL_GRAY, fill_type="solid")
    title_fill  = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    alt_fill    = PatternFill(start_color=ALT_ROW,   end_color=ALT_ROW,   fill_type="solid")
    thin        = Side(border_style="thin", color=BORDER_COL)
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in workbook.worksheets:
        title = titles.get(ws.title, ws.title)
        max_col = ws.max_column or 1
        # Title row
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        c = ws.cell(1, 1, value=title)
        c.font = Font(bold=True, size=14, color=LIGHT_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = title_fill
        ws.row_dimensions[1].height = 24

        # Header styling (row 2)
        if ws.max_row >= 2:
            row2 = ws[2]
            for cell in row2:
                cell.font = Font(bold=True, color=HEADER_TXT)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        # Body styling fast pass
        max_row = ws.max_row
        for i in range(3, max_row + 1, 2):  # zebra banding every other row
            for cell in ws[i]:
                cell.fill = alt_fill

        # Numeric & borders
        numfmt = "#,##0"
        for row in ws.iter_rows(min_row=3, max_row=max_row):
            for cell in row:
                v = cell.value
                if isinstance(v, (int, float)):
                    cell.number_format = numfmt
                cell.border = border

        # Bold first column labels (often AY)
        for i in range(3, max_row + 1):
            ws.cell(i, 1).font = Font(bold=True, color=HEADER_TXT)

        # Orange "Total"
        for i in range(3, max_row + 1):
            if str(ws.cell(i, 1).value).strip().lower() == "total":
                ws.cell(i, 1).font = Font(bold=True, color=ORANGE_ACC)

        # Auto column widths
        col_max = [10] * max_col
        for r in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for j, cell in enumerate(r, start=1):
                val = cell.value
                if val is None:
                    continue
                s = str(val)
                if len(s) > col_max[j-1]:
                    col_max[j-1] = len(s)
        for j, width in enumerate(col_max, start=1):
            ws.column_dimensions[get_column_letter(j)].width = min(max(10, width + 2), 42)

# =========================
# EXCEL I/O (CACHED READERS)
# =========================
@st.cache_data
def _read_excel_sheets(file_bytes: bytes) -> list[str]:
    xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    return xls.sheet_names

@st.cache_data
def _read_sheet(file_bytes: bytes, sheet_name: str, index_col=0) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, engine="openpyxl")
    if index_col is not None and 0 <= int(index_col) < df.shape[1]:
        idx_series = df.columns[int(index_col)]
        df = df.set_index(idx_series)
    return df

# =========================
# PRIMES READER & VALIDATOR
# =========================
def _read_and_validate_primes(file_bytes: bytes) -> pd.DataFrame | None:
    """
    Read primes Excel file and return a DataFrame with:
    - Index: year (int)
    - Column: Premium (numeric)
    Returns None if invalid or empty.
    """
    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
        sheet = xls.sheet_names[0]  # Use first sheet
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, engine="openpyxl")
        
        if df.empty or df.shape[1] < 2:
            return None
        
        # Assume first column is year, second is premium
        year_col = df.columns[0]
        premium_col = df.columns[1]
        
        # Extract and coerce to numeric
        primes = pd.DataFrame({
            "year": pd.to_numeric(df[year_col], errors="coerce"),
            "premium": pd.to_numeric(df[premium_col], errors="coerce")
        })
        
        # Remove rows with NaN in either column
        primes = primes.dropna()
        
        if primes.empty:
            return None
        
        # Convert year to int and set as index
        primes["year"] = primes["year"].astype(int)
        primes = primes.set_index("year")
        primes = primes.rename(columns={"premium": "Premium"})
        
        return primes
    except Exception as e:
        st.warning(f"Failed to read primes file: {e}")
        return None

# =========================
# EXCEL WRITING UTILITIES
# =========================
def _write_final_report_block(xw, sheet_name: str, df_orig: pd.DataFrame, completed_best: pd.DataFrame, ibnr_label: str):
    # Clean & align once
    base_clean = _coerce_age_columns(_coerce_index_to_int(completed_best))
    df_true_clean = _coerce_age_columns(_coerce_index_to_int(df_orig))

    ultimates = base_clean.iloc[:, -1].rename("Ultimate")
    # FAST latest observed per row: forward fill then take last column
    latest_obs = df_true_clean.ffill(axis=1).iloc[:, -1].rename("Latest_Observed")
    ibnr = (ultimates - latest_obs).rename(ibnr_label)

    base_out = _ceil_numeric(base_clean)
    side_table = _ceil_numeric(pd.concat([ultimates, ibnr], axis=1))
    totals = _ceil_numeric(pd.DataFrame({"Ultimate": [float(np.nansum(ultimates.values))],
                                         ibnr_label: [float(np.nansum(ibnr.values))]},
                                        index=["Total"]))

    # Write tables starting row=2 to reserve row=1 for the sheet title
    base_out.to_excel(xw, sheet_name=sheet_name, startrow=1, startcol=0)
    startcol_vectors = base_out.shape[1] + 2
    side_table.to_excel(xw, sheet_name=sheet_name, startrow=1, startcol=startcol_vectors)
    totals.to_excel(xw, sheet_name=sheet_name, startrow=len(side_table) + 3, startcol=startcol_vectors)

# =========================
# REPORT BUILDER
# =========================
def _build_excel_report(
    df_orig: pd.DataFrame,
    completed_best: pd.DataFrame,
    *,
    triangle_type: str,
    ibnr_label: str,
    chosen_method: str | None,
    ultime_start_year: int,
    cuts: list[int],
    selected_methods: list[str],
    ultime_artifacts: dict[int, dict] | None,
    method_scores: pd.DataFrame | None,
    chosen_ay_for_selection: int | None,
    primes_df: pd.DataFrame | None = None,
) -> bytes:
    titles: dict[str, str] = {}
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as xw:
        # --- Meta
        meta_rows = [
            {"Key": "Mode", "Value": "Ultime"},
            {"Key": "Triangle Type", "Value": triangle_type},
        ]
        if chosen_method:
            meta_rows.append({"Key": "Chosen Method", "Value": chosen_method})
        if ultime_start_year is not None:
            meta_rows.append({"Key": "Ultime Start Year", "Value": int(ultime_start_year)})
        if chosen_ay_for_selection is not None:
            meta_rows.append({"Key": "Selection AY", "Value": int(chosen_ay_for_selection)})
        _ceil_numeric(pd.DataFrame(meta_rows)).to_excel(xw, sheet_name="Meta", index=False, startrow=1)
        titles["Meta"] = "Meta • Ultime Mode"
        # --- Completed Last Year + Ultimates + IBNR
        try:
            sheet_name = "Final_Completion"
            base_clean = _coerce_age_columns(_coerce_index_to_int(df_orig))
            comp_clean = _coerce_age_columns(_coerce_index_to_int(completed_best))

            # Get the last AY (latest year) from the original table
            last_year = int(base_clean.index.max())

            # Extract the completed last year row
            last_year_row = comp_clean.loc[[last_year]]

            # Determine label based on triangle type
            ibnr_col_name = "SAP+IBNR" if triangle_type == "Tableau des charges" else "IBNR"

            # Compute ultimate vector and IBNR for each AY
            ultimates = comp_clean.iloc[:, -1].rename("Ultimate")
            last_observed = base_clean.ffill(axis=1).iloc[:, -1].rename("Last_Recorded")
            ibnr = (ultimates - last_observed).rename(ibnr_col_name)

            summary = pd.concat([ultimates, last_observed, ibnr], axis=1)
            totals = pd.DataFrame({"Ultimate": [ultimates.sum()],
                                   "Last_Recorded": [last_observed.sum()],
                                   ibnr_col_name: [ibnr.sum()]},
                                  index=["Total"])

            # Write both sections
            _ceil_numeric(last_year_row).to_excel(xw, sheet_name=sheet_name, startrow=1, startcol=0)
            _ceil_numeric(summary).to_excel(xw, sheet_name=sheet_name, startrow=last_year_row.shape[0] + 4, startcol=0)
            _ceil_numeric(totals).to_excel(xw, sheet_name=sheet_name, startrow=last_year_row.shape[0] + summary.shape[0] + 7, startcol=0)

            titles[sheet_name] = f"Final Completion ({chosen_method}) • Ultimates & {ibnr_col_name}"
        except Exception as e:
            pd.DataFrame({"Error": [str(e)]}).to_excel(xw, sheet_name="Final_Completion", startrow=1)
            titles["Final_Completion"] = "Final Completion (error)"

        # --- Fully Completed Triangle (Final Year)
        try:
            sheet_name = "Final_Completed_Triangle"

            full_completed = _coerce_age_columns(_coerce_index_to_int(completed_best))
            _ceil_numeric(full_completed).to_excel(
                xw, sheet_name=sheet_name, startrow=1, startcol=0
            )

            titles[sheet_name] = f"Full Completed Triangle ({chosen_method}) • Final Year View"
        except Exception as e:
            pd.DataFrame({"Error": [str(e)]}).to_excel(xw, sheet_name="Final_Completed_Triangle", startrow=1)
            titles["Final_Completed_Triangle"] = "Full Completed Triangle (error)"


        # --- Scores
        if method_scores is not None and not method_scores.empty:
            _ceil_numeric(method_scores).to_excel(xw, sheet_name="Method_Scores", index=False, startrow=1)
            titles["Method_Scores"] = "Best Method Selection Scores"

        # --- Per-cut sheets
        if ultime_artifacts:
            for cut_year in cuts:
                pack = ultime_artifacts.get(cut_year, {})
                sheet_name = f"Cut_{cut_year}"
                preds = pack.get("preds", {})

                # Triangle by chosen method (completed at this cut)
                tri_df = None
                if chosen_method and chosen_method in preds:
                    tri_df = preds[chosen_method]
                elif preds:
                    tri_df = preds[sorted(preds.keys())[0]]

                # Ultimes limited to AY ≤ start year, only selected methods that exist
                ult_df = pack.get("ultimes_df", pd.DataFrame())
                if isinstance(ult_df, pd.DataFrame) and not ult_df.empty:
                    keep_methods = [m for m in selected_methods if m in ult_df.columns]
                    ult_tab = ult_df.loc[ult_df.index <= ultime_start_year, keep_methods]
                else:
                    ult_tab = pd.DataFrame()

                if isinstance(tri_df, pd.DataFrame):
                    _ceil_numeric(tri_df).to_excel(xw, sheet_name=sheet_name, index=True, startrow=1, startcol=0)
                    startcol = tri_df.shape[1] + 3
                else:
                    pd.DataFrame({"info": [f"No triangle available for {chosen_method or 'method'} at cut {cut_year}"]}) \
                        .to_excel(xw, sheet_name=sheet_name, index=False, startrow=1, startcol=0)
                    startcol = 4

                if not ult_tab.empty:
                    _ceil_numeric(ult_tab).to_excel(xw, sheet_name=sheet_name, index=True, startrow=1, startcol=startcol)
                else:
                    pd.DataFrame({"info": ["Ultimes not available"]}).to_excel(xw, sheet_name=sheet_name, index=False, startrow=1, startcol=startcol)

                titles[sheet_name] = f"{cut_year}: Triangle ({chosen_method or 'method'}) + Ultimes (AY ≤ {ultime_start_year})"

        # --- Final: per-method grids
        if ultime_artifacts:
            # Determine AY rows (≤ start) once
            grid_ays = set()
            for pack in ultime_artifacts.values():
                udf = pack.get("ultimes_df")
                if isinstance(udf, pd.DataFrame) and not udf.empty:
                    grid_ays.update(int(a) for a in udf.index if int(a) <= int(ultime_start_year))
            grid_rows = sorted(grid_ays)

            for m in selected_methods:
                grid = pd.DataFrame(index=grid_rows, columns=cuts, dtype="float64")
                for cut in cuts:
                    udf = ultime_artifacts.get(cut, {}).get("ultimes_df")
                    if isinstance(udf, pd.DataFrame) and not udf.empty and m in udf.columns:
                        vals = udf[m].reindex(grid_rows)
                        grid.loc[:, cut] = vals.values
                base = "UltimesGrid_"
                sheet_name = f"{base}{m[:31 - len(base)]}"

                _ceil_numeric(grid).to_excel(xw, sheet_name=sheet_name, index=True, startrow=1)
                titles[sheet_name] = f"{m}: Ultimes Grid (AY ≤ {ultime_start_year}; Cuts {cuts[0]}–{cuts[-1]})"

        # --- S/P Ratio Sheet (if primes provided)
        if primes_df is None or primes_df.empty:
            st.info("ℹ️ No valid primes provided — S/P Ratio sheet not generated.")
        elif primes_df is not None and not primes_df.empty:
            try:
                comp_clean = _coerce_age_columns(_coerce_index_to_int(completed_best))
                ultimates = comp_clean.iloc[:, -1].rename("Ultimate")
                
                # Align ultimates with primes by year (convert index to int for matching)
                ultimates_idx = ultimates.index.astype(int)
                primes_idx = primes_df.index.astype(int)
                
                # Find matching years
                matching_years = ultimates_idx.intersection(primes_idx)
                
                if len(matching_years) > 0:
                    sp_data = pd.DataFrame({
                        "Ultimate": ultimates.loc[matching_years],
                        "Premium": primes_df.loc[matching_years, "Premium"]
                    })
                    
                    # Keep only rows where both ultimate and premium exist
                    sp_data = sp_data.dropna()
                    
                    if not sp_data.empty:
                        # Calculate S/P ratio (Sinistre/Prime)
                        sp_data["S/P"] = sp_data["Ultimate"] / sp_data["Premium"]
                        
                        # Format for display
                        sp_display = sp_data.copy()
                        sp_display["Ultimate"] = np.rint(np.ceil(sp_display["Ultimate"])).astype("Int64")
                        sp_display["Premium"] = np.rint(np.ceil(sp_display["Premium"])).astype("Int64")
                        sp_display["S/P"] = sp_display["S/P"].apply(lambda x: f"{x:.4f}" if pd.notna(x) and np.isfinite(x) else "")
                        
                        sp_display.to_excel(xw, sheet_name="S_P_Ratio", index=True, startrow=1)
                        titles["S_P_Ratio"] = f"S/P Ratio (Sinistre/Prime) — Method: {chosen_method}"
            except Exception as e:
                st.warning(f"Could not compute S/P ratio: {e}")

        # Styling
        _apply_global_excel_styles(xw.book, titles)

    buffer.seek(0)
    return buffer.getvalue()

# =========================
# APP UI (Ultime only) — 5-PHASE WORKFLOW
# =========================
_brand_header()

# ============================================================================
# PHASE 1: DATA INTAKE
# ============================================================================
st.markdown('<div class="g3m-phase">', unsafe_allow_html=True)
st.markdown("### Phase 1: Data Intake")
st.markdown('<span class="g3m-muted">Upload and validate your loss triangle</span>', unsafe_allow_html=True)

file = st.file_uploader("Drop your Excel file (.xlsx)", type=["xlsx"], accept_multiple_files=False)

if not file:
    st.info("Upload an Excel file to begin.")
    st.stop()

file_bytes = file.read()
try:
    sheets = _read_excel_sheets(file_bytes)
except Exception as e:
    st.error(f"Couldn't read sheets: {e}")
    st.stop()

with st.container():
    st.markdown('---', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        sheet = st.selectbox("Sheet", options=sheets, index=0)
    with col2:
        triangle_type = st.radio("Type", ["Tableau des charges", "Tableau Cummulative"], index=0, horizontal=True)

ibnr_label = "IBNR" if triangle_type == "Tableau des charges" else "SAP+IBNR"
index_col = 0

try:
    df_raw = _read_sheet(file_bytes, sheet, index_col=index_col)
except Exception as e:
    st.error(f"Failed to read the selected sheet: {e}")
    st.stop()

with st.container():
    st.markdown('---', unsafe_allow_html=True)
    st.markdown("**Data Preview**")
    _preview = df_raw.copy()
    _preview.columns = _preview.columns.astype(str)
    st.dataframe(_format_display(_preview.head(15)), use_container_width=True)

# Determine valid years early
try:
    full = _coerce_age_columns(_coerce_index_to_int(df_raw))
    years = list(map(int, full.index.tolist()))
    min_year = min(years); max_year = max(years)
    valid_starts = [y for y in range(min_year, max_year)]
    if not valid_starts:
        st.error("Not enough years to run Ultime.")
        st.stop()
except Exception as e:
    st.error(f"Unable to parse years: {e}")
    st.stop()

# ============================================================================
# PHASE 2: ASSUMPTIONS VALIDATION
# ============================================================================
st.markdown('<div class="g3m-phase">', unsafe_allow_html=True)
st.markdown("### Phase 2: Assumptions Validation")
st.markdown('<span class="g3m-muted">Verify model conditions are met</span>', unsafe_allow_html=True)

with st.sidebar:
    st.header("Assumption Settings")
    r2_threshold = st.number_input(
        "R² threshold (Factors)",
        min_value=0.0, max_value=1.0, value=0.95, step=0.01,
        help="Higher = stricter stability requirements."
    )
    tail_factor = st.number_input(
        "Tail factor",
        min_value=0.0, value=1.0, step=0.01,
        help="Development extension beyond observed triangle."
    )
    z_alpha_2 = st.number_input(
        "Z (Calendar test)",
        min_value=0.0, value=1.96, step=0.01,
        help="Confidence level for calendar effect test."
    )
    tie_tol = st.number_input(
        "Tie tolerance",
        min_value=0.0, value=0.0, step=0.01
    )

col1, col2 = st.columns(2)
with col1:
    st.markdown('---', unsafe_allow_html=True)
    st.markdown("**Factors Assumption**")
    st.markdown('<span class="g3m-muted">R² stability test</span>', unsafe_allow_html=True)
    if st.button("Validate Factors", key="btn_validate_factors", use_container_width=True):
        with st.spinner("Testing factors R²..."):
            try:
                if AssumptionFactors is None:
                    raise RuntimeError("AssumptionFactors module not found.")
                factors_ok = AssumptionFactors(
                    full,
                    output_path="Correlation_Factors_Output2.xlsx",
                    r2_threshold=float(r2_threshold),
                    tail_factor=float(tail_factor),
                    eps_denom=0.0
                )
                st.session_state.factors_ok = factors_ok
            except Exception as e:
                st.error(f"Test failed: {e}")
                st.session_state.factors_ok = None
    
    if "factors_ok" in st.session_state:
        if st.session_state.factors_ok is True:
            st.markdown('<span class="g3m-success">✓ Accepted (R² ≥ threshold)</span>', unsafe_allow_html=True)
        elif st.session_state.factors_ok is False:
            st.markdown('<span class="g3m-error">✗ Rejected (R² < threshold)</span>', unsafe_allow_html=True)
    else:
        st.markdown('<span class="g3m-muted">Not yet validated</span>', unsafe_allow_html=True)

with col2:
    st.markdown('---', unsafe_allow_html=True)
    st.markdown("**Calendar Assumption**")
    st.markdown('<span class="g3m-muted">Time effect hypothesis test</span>', unsafe_allow_html=True)
    if st.button("Validate Calendar", key="btn_validate_calendar", use_container_width=True):
        with st.spinner("Testing calendar effect..."):
            try:
                if AssumptionCalendar is None:
                    raise RuntimeError("AssumptionCalendar module not found.")
                calendar_ok = AssumptionCalendar(
                    full,
                    output_path="Calendar_Effect_Test2.xlsx",
                    z_alpha_2=float(z_alpha_2),
                    eps_denom=0.0,
                    tie_tol=float(tie_tol)
                )
                st.session_state.calendar_ok = calendar_ok
            except Exception as e:
                st.error(f"Test failed: {e}")
                st.session_state.calendar_ok = None
    
    if "calendar_ok" in st.session_state:
        if st.session_state.calendar_ok is True:
            st.markdown('<span class="g3m-success">✓ Accepted (Z in CI)</span>', unsafe_allow_html=True)
        elif st.session_state.calendar_ok is False:
            st.markdown('<span class="g3m-error">✗ Rejected (Z out CI)</span>', unsafe_allow_html=True)
    else:
        st.markdown('<span class="g3m-muted">Not yet validated</span>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# PHASE 3: METHOD STRATEGY
# ============================================================================
st.markdown('<div class="g3m-phase">', unsafe_allow_html=True)
st.markdown("### Phase 3: Method Strategy")
st.markdown('<span class="g3m-muted">Select completion methods for analysis</span>', unsafe_allow_html=True)

def render_methods_ui(default_selected: list[str]) -> tuple[list[str], dict]:
    configs = st.session_state.get("METHOD_CONFIGS", {})
    selected: list[str] = []

    for name in ALL_METHOD_NAMES:
        if not callable(globals().get(name, None)):
            continue
        meta = METHOD_META.get(name, {"title": name, "desc": ""})
        st.markdown('---', unsafe_allow_html=True)
        cols = st.columns([0.75, 0.25])
        with cols[0]:
            st.markdown(f"**{meta['title']}**")
            st.markdown(f"<span class='g3m-muted'>{meta.get('desc', '')}</span>", unsafe_allow_html=True)
            pick = st.checkbox("Select", key=f"pick_{name}", value=(name in default_selected))
        with cols[1]:
            if name in ("Poissonbootstrap", "ChainLadderBootstrap"):
                with st.expander("Settings", expanded=False):
                    B_default = int(configs.get(name, {}).get("B", 200))
                    include_default = bool(configs.get(name, {}).get("Report", False))
                    rname_default = configs.get(name, {}).get(
                        "report_name",
                        "Poisson_Bootstrap_Report" if name == "Poissonbootstrap" else "CL_Bootstrap_Report"
                    )
                    B = st.number_input("Simulations", min_value=100, max_value=10000, step=100, value=B_default, key=f"B_{name}")
                    include = st.checkbox("Include report", value=include_default, key=f"rep_{name}")
                    rname = st.text_input("Name", value=rname_default, key=f"name_{name}")
                    configs[name] = {"B": int(B), "Report": include, "report_name": rname}

        if pick:
            selected.append(name)

    st.session_state.METHOD_CONFIGS = configs
    return selected, configs

selected_methods, _method_cfgs = render_methods_ui([m for m in ALL_METHOD_NAMES if callable(globals().get(m, None))])
st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# PHASE 4: STABILITY & SELECTION
# ============================================================================
st.markdown('<div class="g3m-phase">', unsafe_allow_html=True)
st.markdown("### Phase 4: Ultime Stability & Method Selection")
st.markdown('<span class="g3m-muted">Analyze consistency across cut years</span>', unsafe_allow_html=True)

with st.sidebar:
    st.header("Ultime Settings")
    ultime_start_year = st.selectbox(
        "Start year for ultime analysis",
        options=valid_starts,
        index=max(0, len(valid_starts)//2)
    )
    ultime_decay_r = st.number_input(
        "History decay (0<r≤1)",
        min_value=0.01, max_value=1.0, value=0.85, step=0.01,
        help="Weight on recent cuts vs historical."
    )
    selection_ay_weight = st.number_input(
        "Closeness vs Stability",
        min_value=0.0, max_value=1.0, value=0.70, step=0.05,
        help="0=stability only, 1=closeness only."
    )

selection_ay = ultime_start_year

run_clicked = st.button("Run Ultime Stability Analysis", type="primary", disabled=(len(selected_methods) == 0), use_container_width=True)

best_method = None
method_scores = pd.DataFrame()
ultime_artifacts: dict[int, dict] = {}
sequences = {}
final_ultimes = pd.DataFrame()
cuts: list[int] = []

if run_clicked:
    with st.spinner(f"Running Ultime stability analysis from {ultime_start_year} onwards..."):
        try:
            sequences, ultime_artifacts, final_ultimes, cuts = _cached_cached_ultime_cycle(
                df_raw, selected_methods, int(ultime_start_year), decay_r=float(ultime_decay_r)
            )
            best_method, method_scores = _select_best_method(
                sequences, cuts, int(selection_ay),
                closeness_weight=float(selection_ay_weight),
                decay_r=float(ultime_decay_r)
            )
        except Exception as e:
            st.error(f"Ultime analysis failed: {e}")
            st.stop()

    if not method_scores.empty:
        st.markdown('---', unsafe_allow_html=True)
        st.markdown("**Method Comparison Scores**")
        st.markdown('<span class="g3m-muted">Lower scores indicate better stability and consistency</span>', unsafe_allow_html=True)
        st.dataframe(_format_display(method_scores), use_container_width=True)

    if best_method:
        st.markdown('---', unsafe_allow_html=True)
        st.markdown(f'<span class="g3m-success">✓ Recommended: <strong>{best_method}</strong> for AY {selection_ay}</span>', unsafe_allow_html=True)
    else:
        st.info("No clear best method; select manually in Phase 5.")

st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# PHASE 5: REPORT GENERATION
# ============================================================================
st.markdown('<div class="g3m-phase">', unsafe_allow_html=True)
st.markdown("### Phase 5: Report Generation")
st.markdown('<span class="g3m-muted">Create and download finalized Excel report</span>', unsafe_allow_html=True)

# Optional primes file for S/P ratio
primes_file = st.file_uploader("(Optional) Primes file for S/P ratio (.xlsx)", type=["xlsx"], accept_multiple_files=False, key="primes_upload")
if primes_file is not None:
    try:
        primes_bytes = primes_file.read()
        if primes_bytes:
            primes_data = _read_and_validate_primes(primes_bytes)
            if primes_data is not None and not primes_data.empty:
                st.session_state.PRIMES_DATA = primes_data
                st.success("✓ Primes file loaded")
            else:
                st.warning("Primes file is empty or invalid")
    except Exception as e:
        st.warning(f"Could not read primes: {e}")

st.markdown('---', unsafe_allow_html=True)
st.markdown("**Select Completion Method**")
options_for_completion = selected_methods if selected_methods else [m for m in ALL_METHOD_NAMES if callable(globals().get(m, None))]

if "chosen_method" not in st.session_state:
    st.session_state.chosen_method = None
if "chosen_manually_set" not in st.session_state:
    st.session_state.chosen_manually_set = False

if best_method and not st.session_state.chosen_manually_set:
    st.session_state.chosen_method = best_method if best_method in options_for_completion else (options_for_completion[0] if options_for_completion else None)

def _mark_manual_choice():
    st.session_state.chosen_manually_set = True

col1, col2 = st.columns([2, 1])
with col1:
    chosen = st.selectbox(
        "Method",
        options=options_for_completion,
        index=options_for_completion.index(st.session_state.chosen_method) if st.session_state.chosen_method in options_for_completion else 0,
        key="chosen_method_final",
        on_change=_mark_manual_choice,
    )
with col2:
    report_name = st.text_input("Filename", value="Ultime_Report", label_visibility="collapsed")

build_clicked = st.button("Generate Report", type="primary", use_container_width=True)

if build_clicked:
    with st.spinner("Generating Excel report..."):
        try:
            func = globals().get(st.session_state.chosen_method or chosen, None)
            if not callable(func):
                raise ValueError(f"Selected method is not callable.")
            base = _coerce_age_columns(_coerce_index_to_int(df_raw))
            base_for_method = base

            with pd.option_context("mode.copy_on_write", False):
                out = _call_with_signature(
                    func,
                    _make_writable_df(base_for_method),
                    st.session_state.get("METHOD_CONFIGS", {}).get(st.session_state.chosen_method or chosen, {})
                )

            if isinstance(out, tuple):
                pred_full_raw, report_bytes = out
                if report_bytes is not None:
                    st.session_state.setdefault("BOOTSTRAP_REPORTS", {})
                    st.session_state.BOOTSTRAP_REPORTS[st.session_state.chosen_method or chosen] = report_bytes
            else:
                pred_full_raw = out

            pred_full = _coerce_age_columns(_coerce_index_to_int(pred_full_raw)).reindex(index=base.index, columns=base.columns)
            completed = base.copy()
            mask = completed.isna() & pred_full.notna()
            completed[mask] = pred_full[mask]
            completed = completed.ffill(axis=1)
        except Exception as e:
            st.error(f"Completion failed: {e}")
            st.stop()

        try:
            if not cuts or not ultime_artifacts:
                sequences, ultime_artifacts, final_ultimes, cuts = _cached_cached_ultime_cycle(
                    df_raw, selected_methods, int(ultime_start_year), decay_r=float(ultime_decay_r)
                )
            if method_scores is None or method_scores.empty:
                _, method_scores = _select_best_method(
                    sequences, cuts, int(selection_ay),
                    closeness_weight=float(selection_ay_weight),
                    decay_r=float(ultime_decay_r)
                )
        except Exception as e:
            st.error(f"Pre-build analysis failed: {e}")
            st.stop()

        try:
            xlsx_bytes = _build_excel_report(
                df_orig=df_raw,
                completed_best=completed,
                triangle_type=triangle_type,
                ibnr_label=ibnr_label,
                chosen_method=st.session_state.chosen_method or chosen,
                ultime_start_year=int(ultime_start_year),
                cuts=cuts,
                selected_methods=selected_methods,
                ultime_artifacts=ultime_artifacts,
                method_scores=method_scores,
                chosen_ay_for_selection=int(selection_ay),
                primes_df=st.session_state.get("PRIMES_DATA"),
            )
        except Exception as e:
            st.error(f"Report generation failed: {e}")
            st.stop()

        st.markdown('---', unsafe_allow_html=True)
        st.markdown("**Report Ready**")
        st.markdown(f"✓ Completed with <strong>{st.session_state.chosen_method or chosen}</strong>")

        download_base = _sanitize_filename(report_name or "Ultime_Report")
        st.download_button(
            label="Download Excel Report",
            data=xlsx_bytes,
            file_name=f"{download_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )

        bootstrap_reports = st.session_state.get("BOOTSTRAP_REPORTS", {})
        if bootstrap_reports:
            st.markdown("---")
            st.markdown("**Bootstrap Reports**")
            if "Poissonbootstrap" in bootstrap_reports and bootstrap_reports["Poissonbootstrap"] is not None:
                st.download_button(
                    "Download Poisson Bootstrap Report",
                    data=bootstrap_reports["Poissonbootstrap"],
                    file_name="Poisson_Bootstrap_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="poisson_dl",
                    use_container_width=True
                )
            if "ChainLadderBootstrap" in bootstrap_reports and bootstrap_reports["ChainLadderBootstrap"] is not None:
                st.download_button(
                    "Download Chain Ladder Bootstrap Report",
                    data=bootstrap_reports["ChainLadderBootstrap"],
                    file_name="ChainLadder_Bootstrap_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="cl_dl",
                    use_container_width=True
                )

st.markdown('</div>', unsafe_allow_html=True)
